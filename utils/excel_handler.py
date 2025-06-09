import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from utils.xbrl_handler import *  # NOQA


def export_to_excel_template(df: pd.DataFrame,
                             sheet_name: str,
                             workbook: openpyxl.Workbook = None,
                             workbook_path: str = None):
    df = df.copy()
    if workbook is None:
        if workbook_path is None:
            raise ValueError(
                "Either workbook or workbook_path must be provided.")
        # Load the workbook from the path if no workbook object is given
        bs_wb = openpyxl.load_workbook(workbook_path)
    else:
        # Use the workbook object that was passed in
        bs_wb = workbook

    ws = bs_wb[sheet_name]
    # ws

    a_col_len = len(ws['A'])
    coord_dict = dict()
    bold_title = []

    for i in range(a_col_len):
        row_number = i + 1
        coord = f"A{row_number}"
        cell = ws[coord]
        cell_value = cell.value

        if cell.font.bold:
            bold_title.append(row_number)

        coord_dict[cell_value] = row_number

    # bs_df = bs_df.replace({pd.NA: np.nan})
    df.drop('name', axis=1, inplace=True)
    df['excel_row'] = df['formatted_name'].map(coord_dict)

    for index, row in df.iterrows():
        # print(row)
        formatted_name = row.pop('formatted_name')  # NOQA
        # print(f"Filling {formatted_name}...")
        to_be_filled_row = row.pop('excel_row')
        to_be_filled_value = row.to_list()
        to_be_filled_year = row.index.to_list()

        all_columns = ws[f'{to_be_filled_row}']

        first_blank_column = None
        for column in all_columns:
            col_value = column.value
            if col_value is None:
                first_blank_column = column.column
                break

        # FILL YEAR
        for i, value in enumerate(to_be_filled_year):
            # print(value)
            # column_char =
            if i == 0:
                cell_to_be_filled = get_column_letter(first_blank_column + 1)
            else:
                current_to_be_filled = first_blank_column + i + 1
                cell_to_be_filled = get_column_letter(current_to_be_filled)

            cell_cord = f"{cell_to_be_filled}3"
            # print(f"Filling cell {cell_cord} with {value}")
            ws[cell_cord] = value

        # FILL VALUE
        for i, value in enumerate(to_be_filled_value):
            # print(i)
            # column_char =
            if i == 0:
                cell_to_be_filled = get_column_letter(first_blank_column + 1)
            else:
                current_to_be_filled = first_blank_column + i + 1
                cell_to_be_filled = get_column_letter(current_to_be_filled)

            cell_cord = f"{cell_to_be_filled}{to_be_filled_row}"
            # print(f"Filling cell {cell_cord} with {value}")

            if pd.isna(value):
                value = np.nan
            ws[cell_cord] = value

    range_a_col = list(range(1, a_col_len + 1))
    # range_a_col
    not_bold = [x for x in range_a_col if x not in bold_title]
    # exclusion

    # i = exclusion[0]

    empty_dict = dict()
    for i in not_bold:
        all_rows = ws[i]
        # print(f"Examining row {i}")
        rows_value = []
        for idx, row in enumerate(all_rows):
            # print(f"idx: {idx}")
            if idx != 0:
                value = row.value
                # print(value)
                rows_value.append(value)
        empty_dict[i] = rows_value

    to_be_hide_row = []
    for k, v in empty_dict.items():
        all_empty = all(
            x is None or (
                isinstance(
                    x,
                    float) and np.isnan(x)) for x in v)
        if all_empty:
            to_be_hide_row.append(k)

    for row in to_be_hide_row:
        ws.row_dimensions[row].hidden = True

    # Freeze PAnes
    ws.views.sheetView[0].topLeftCell = 'A1'  # scroll to top left corner
    ws.freeze_panes = 'C4'
    # ws.freeze_panes = 'G4'

    return bs_wb


def export_general_report_to_excel(stock_code: str,
                                   df: pd.DataFrame,
                                   workbook_path: str,
                                   id_availability_dict: dict,
                                   quarterly: bool,
                                   convert_billion_idr: bool,
                                   convert_million_usd: bool):

    gi_df = get_general_financial_table(
        combined_df=df,
        name_mapping_path='./schema/all_general_information_1000000.csv',
        table_type="general_info",
        convert_billion_idr=convert_billion_idr,
        convert_million_usd=convert_million_usd)

    processed_wb = export_to_excel_template(
        df=gi_df,
        workbook_path=workbook_path,
        sheet_name='GENERAL INFO')

    if id_availability_dict['1210000']:
        bs_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_balance_sheet_1210000.csv',
            table_type="balance_sheet",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=bs_df,
                                                workbook=processed_wb,
                                                sheet_name='BALANCE SHEET')
    else:
        processed_wb.remove(processed_wb['BALANCE SHEET'])

    if id_availability_dict['1311000'] or id_availability_dict['1321000'] or id_availability_dict['1312000']:
        is_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_income_statement_1321000.csv',
            table_type="income_statement",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=is_df,
                                                workbook=processed_wb,
                                                sheet_name='INCOME STATEMENT')

        if quarterly:
            is_quarterly_df = convert_cumulative_to_quarterly(is_df)

            processed_wb = export_to_excel_template(
                df=is_quarterly_df,
                workbook=processed_wb,
                sheet_name='INCOME STATEMENT QoQ')
        else:
            processed_wb.remove(processed_wb['INCOME STATEMENT QoQ'])

    else:
        processed_wb.remove(processed_wb['INCOME STATEMENT'])
        processed_wb.remove(processed_wb['INCOME STATEMENT QoQ'])

    if id_availability_dict['1510000']:
        cf_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_cash_flow_1510000.csv',
            table_type="cash_flow",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=cf_df,
                                                workbook=processed_wb,
                                                sheet_name='CASH FLOW')

        if quarterly:
            cf_quarterly_df = convert_cumulative_to_quarterly(cf_df)

            processed_wb = export_to_excel_template(df=cf_quarterly_df,
                                                    workbook=processed_wb,
                                                    sheet_name='CASH FLOW QoQ')
        else:
            processed_wb.remove(processed_wb['CASH FLOW QoQ'])

    else:
        processed_wb.remove(processed_wb['CASH FLOW'])
        processed_wb.remove(processed_wb['CASH FLOW QoQ'])

    if id_availability_dict['1610000']:
        ap_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_accounting_policy_1610000.csv',
            table_type="accounting_policies",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='ACCOUNTING POLICIES')
    else:
        processed_wb.remove(processed_wb['ACCOUNTING POLICIES'])

    if id_availability_dict['1630000']:
        inventory_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_inventory_1630000.csv',
            table_type="inventory_breakdown",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=inventory_df,
            workbook=processed_wb,
            sheet_name='INVENTORY BREAKDOWN')
    else:
        processed_wb.remove(processed_wb['INVENTORY BREAKDOWN'])

    if id_availability_dict['1632000']:
        inventory_notes_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_inventory_notes_1632000.csv',
            table_type="inventory_notes",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=inventory_notes_df,
                                                workbook=processed_wb,
                                                sheet_name='INVENTORY NOTES')
    else:
        processed_wb.remove(processed_wb['INVENTORY NOTES'])

    if id_availability_dict['1620100']:
        receivable_currency_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_receivable_by_currency_1620100.csv',
            table_type="receivable_by_currency",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=receivable_currency_df,
            workbook=processed_wb,
            sheet_name='RECEIVABLE BY CURRENCY')
    else:
        processed_wb.remove(processed_wb['RECEIVABLE BY CURRENCY'])

    if id_availability_dict['1620200']:
        receivable_aging_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_receivable_by_aging_1620200.csv',
            table_type="receivable_by_aging",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=receivable_aging_df,
            workbook=processed_wb,
            sheet_name='RECEIVABLE BY AGING')
    else:
        processed_wb.remove(processed_wb['RECEIVABLE BY AGING'])

    if id_availability_dict['1620300']:
        receivable_parties_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_receivable_by_parties_1620300.csv',
            table_type="receivable_by_parties",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=receivable_parties_df,
            workbook=processed_wb,
            sheet_name='RECEIVABLE BY PARTIES')
    else:
        processed_wb.remove(processed_wb['RECEIVABLE BY PARTIES'])

    if id_availability_dict['1620500']:
        receivable_allowance_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_receivable_allowances_1620500.csv',
            table_type="receivable_allowances",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=receivable_allowance_df,
            workbook=processed_wb,
            sheet_name='RECEIVABLE ALLOWANCES')
    else:
        processed_wb.remove(processed_wb['RECEIVABLE ALLOWANCES'])

    if id_availability_dict['1620400']:
        receivable_area_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_receivable_by_area_1620400.csv',
            table_type="receivable_by_area",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=receivable_area_df,
            workbook=processed_wb,
            sheet_name='RECEIVABLE BY AREA')
    else:
        processed_wb.remove(processed_wb['RECEIVABLE BY AREA'])

    if id_availability_dict['1640100']:
        payable_currency_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_payable_by_currency_1640100.csv',
            table_type="payable_by_currency",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=payable_currency_df,
            workbook=processed_wb,
            sheet_name='PAYABLE BY CURRENCY')
    else:
        processed_wb.remove(processed_wb['PAYABLE BY CURRENCY'])

    if id_availability_dict['1640200']:
        payable_aging_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_payable_by_aging_1640200.csv',
            table_type="payable_by_aging",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=payable_aging_df,
                                                workbook=processed_wb,
                                                sheet_name='PAYABLE BY AGING')
    else:
        processed_wb.remove(processed_wb['PAYABLE BY AGING'])

    if id_availability_dict['1640300']:
        payable_parties_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_payable_by_parties_1640300.csv',
            table_type="payable_by_parties",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=payable_parties_df,
            workbook=processed_wb,
            sheet_name='PAYABLE BY PARTIES')
    else:
        processed_wb.remove(processed_wb['PAYABLE BY PARTIES'])

    if id_availability_dict['1691000a']:
        lt_bank_loan_value_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_lt_bank_loans_value_169100a.csv',
            table_type="long_term_bank_loan_value",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=lt_bank_loan_value_df,
            workbook=processed_wb,
            sheet_name='LONG TERM BANK LOAN VALUE')
    else:
        processed_wb.remove(processed_wb['LONG TERM BANK LOAN VALUE'])

    if id_availability_dict['1691100']:
        lt_bank_loan_notes_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_lt_bank_loan_notes_1691100.csv',
            table_type="long_term_bank_loan_notes",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=lt_bank_loan_notes_df,
            workbook=processed_wb,
            sheet_name='LONG TERM BANK LOAN NOTES')
    else:
        processed_wb.remove(processed_wb['LONG TERM BANK LOAN NOTES'])

    if id_availability_dict['1692000']:
        lt_bank_loan_interest_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_lt_bank_interest_1692000.csv',
            table_type="long_term_bank_loan_interest",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=lt_bank_loan_interest_df,
            workbook=processed_wb,
            sheet_name='LONG TERM BANK INTEREST')
    else:
        processed_wb.remove(processed_wb['LONG TERM BANK INTEREST'])

    if id_availability_dict['1693000']:
        lt_bank_loan_value_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_st_bank_loans_value_1693000.csv',
            table_type="short_term_bank_loan_value",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=lt_bank_loan_value_df,
            workbook=processed_wb,
            sheet_name='SHORT TERM BANK LOAN VALUE')
    else:
        processed_wb.remove(processed_wb['SHORT TERM BANK LOAN VALUE'])

    if id_availability_dict['1693100']:
        st_bank_loan_notes_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_st_bank_loan_notes_1693100.csv',
            table_type="short_term_bank_loan_notes",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=st_bank_loan_notes_df,
            workbook=processed_wb,
            sheet_name='SHORT TERM BANK LOAN NOTES')
    else:
        processed_wb.remove(processed_wb['SHORT TERM BANK LOAN NOTES'])

    if id_availability_dict['1696000']:
        st_bank_loan_notes_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_st_bank_interest_1696000.csv',
            table_type="short_term_bank_loan_interest",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=st_bank_loan_notes_df,
            workbook=processed_wb,
            sheet_name='SHORT TERM BANK INTEREST')
    else:
        processed_wb.remove(processed_wb['SHORT TERM BANK INTEREST'])

    if id_availability_dict['1616000']:
        rev_by_parties_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_revenue_by_parties_1616000.csv',
            table_type="revenue_by_parties",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=rev_by_parties_df,
            workbook=processed_wb,
            sheet_name='REVENUE BY PARTIES')

        if quarterly:
            rev_by_parties_quarterly_df = convert_cumulative_to_quarterly(
                rev_by_parties_df)

            processed_wb = export_to_excel_template(
                df=rev_by_parties_quarterly_df,
                workbook=processed_wb,
                sheet_name='REVENUE BY PARTIES QoQ')
        else:
            processed_wb.remove(processed_wb['REVENUE BY PARTIES QoQ'])
    else:
        processed_wb.remove(processed_wb['REVENUE BY PARTIES'])
        processed_wb.remove(processed_wb['REVENUE BY PARTIES QoQ'])

    if id_availability_dict['1617000']:
        rev_by_type_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_revenue_by_type_1617000.csv',
            table_type="revenue_by_type",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=rev_by_type_df,
                                                workbook=processed_wb,
                                                sheet_name='REVENUE BY TYPE')

        if quarterly:
            rev_by_type_quarterly_df = convert_cumulative_to_quarterly(
                rev_by_type_df)

            processed_wb = export_to_excel_template(
                df=rev_by_type_quarterly_df,
                workbook=processed_wb,
                sheet_name='REVENUE BY TYPE QoQ')
        else:
            processed_wb.remove(processed_wb['REVENUE BY TYPE QoQ'])
    else:
        processed_wb.remove(processed_wb['REVENUE BY TYPE'])
        processed_wb.remove(processed_wb['REVENUE BY TYPE QoQ'])

    if id_availability_dict['1618000']:
        rev_by_sources_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_revenue_by_sources_1618000.csv',
            table_type="revenue_by_source",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=rev_by_sources_df,
            workbook=processed_wb,
            sheet_name='REVENUE BY SOURCES')

        if quarterly:
            rev_by_sources_quarterly_df = convert_cumulative_to_quarterly(
                rev_by_sources_df)

            processed_wb = export_to_excel_template(
                df=rev_by_sources_quarterly_df,
                workbook=processed_wb,
                sheet_name='REVENUE BY SOURCES QoQ')
        else:
           processed_wb.remove(processed_wb['REVENUE BY SOURCES QoQ']) 
    else:
        processed_wb.remove(processed_wb['REVENUE BY SOURCES'])
        processed_wb.remove(processed_wb['REVENUE BY SOURCES QoQ'])

    if id_availability_dict['1619000']:
        rev_morethan_10percent_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_revenue_10percent_1619000.csv',
            table_type="revenue_morethan_10percent",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=rev_morethan_10percent_df,
                                                workbook=processed_wb,
                                                sheet_name='REVENUE >10%')

        if quarterly:
            rev_morethan_10percent_quarterly_df = convert_cumulative_to_quarterly(
                rev_morethan_10percent_df)

            processed_wb = export_to_excel_template(
                df=rev_morethan_10percent_quarterly_df,
                workbook=processed_wb,
                sheet_name='REVENUE >10% QoQ')
        else:
           processed_wb.remove(processed_wb['REVENUE >10% QoQ']) 
    else:
        processed_wb.remove(processed_wb['REVENUE >10%'])
        processed_wb.remove(processed_wb['REVENUE >10% QoQ'])

    if id_availability_dict['1670000']:
        cogs_breakdown_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_cogs_1670000.csv',
            table_type="cogs_breakdown",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=cogs_breakdown_df,
                                                workbook=processed_wb,
                                                sheet_name='COGS BREAKDOWN')
        if quarterly:
            cogs_breakdown_quarterly_df = convert_cumulative_to_quarterly(
                cogs_breakdown_df)

            processed_wb = export_to_excel_template(
                df=cogs_breakdown_quarterly_df,
                workbook=processed_wb,
                sheet_name='COGS BREAKDOWN QoQ')
        else:
          processed_wb.remove(processed_wb['COGS BREAKDOWN QoQ'])  
    else:
        processed_wb.remove(processed_wb['COGS BREAKDOWN'])
        processed_wb.remove(processed_wb['COGS BREAKDOWN QoQ'])

    if id_availability_dict['1671000']:
        cogs_notes_df = get_general_financial_table(
            combined_df=df,
            name_mapping_path='./schema/general_cogs_notes_1671000.csv',
            table_type="cogs_notes",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=cogs_notes_df,
                                                workbook=processed_wb,
                                                sheet_name='COGS NOTES')
    else:
        processed_wb.remove(processed_wb['COGS NOTES'])

    if quarterly:
        processed_wb.save(
            f"./result/{stock_code}_Financial_Statement_Quarterly.xlsx")
    else:
        processed_wb.save(f"./result/{stock_code}_Financial_Statement.xlsx")


def export_finance_report_to_excel(stock_code: str,
                                   df: pd.DataFrame,
                                   workbook_path: str,
                                   id_availability_dict: dict,
                                   quarterly: bool,
                                   convert_billion_idr: bool,
                                   convert_million_usd: bool):

    gi_df = get_finance_financial_table(
        combined_df=df,
        name_mapping_path='./schema/all_general_information_1000000.csv',
        table_type="general_info",
        convert_billion_idr=convert_billion_idr,
        convert_million_usd=convert_million_usd)

    processed_wb = export_to_excel_template(
        df=gi_df,
        workbook_path=workbook_path,
        sheet_name='GENERAL INFO')

    if id_availability_dict['4220000']:
        bs_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_balance_sheet_4220000.csv',
            table_type="balance_sheet",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=bs_df,
                                                workbook=processed_wb,
                                                sheet_name='BALANCE SHEET')
    else:
        processed_wb.remove(processed_wb['BALANCE SHEET'])

    if id_availability_dict['4312000'] or id_availability_dict['4322000'] or id_availability_dict['1312000']:
        is_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_income_statement_4312000.csv',
            table_type="income_statement",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=is_df,
                                                workbook=processed_wb,
                                                sheet_name='INCOME STATEMENT')

        if quarterly:
            is_quarterly_df = convert_cumulative_to_quarterly(is_df)

            processed_wb = export_to_excel_template(
                df=is_quarterly_df,
                workbook=processed_wb,
                sheet_name='INCOME STATEMENT QoQ')
        else:
            processed_wb.remove(processed_wb['INCOME STATEMENT QoQ'])

    else:
        processed_wb.remove(processed_wb['INCOME STATEMENT'])
        processed_wb.remove(processed_wb['INCOME STATEMENT QoQ'])

    if id_availability_dict['4510000']:
        cf_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_cash_flow_4510000.csv',
            table_type="cash_flow",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=cf_df,
                                                workbook=processed_wb,
                                                sheet_name='CASH FLOW')
        if quarterly:
            cf_quarterly_df = convert_cumulative_to_quarterly(cf_df)

            processed_wb = export_to_excel_template(df=cf_quarterly_df,
                                                    workbook=processed_wb,
                                                    sheet_name='CASH FLOW QoQ')
        else:
            processed_wb.remove(processed_wb['CASH FLOW QoQ'])
    else:
        processed_wb.remove(processed_wb['CASH FLOW'])
        processed_wb.remove(processed_wb['CASH FLOW QoQ'])

    if id_availability_dict['4610000']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_accounting_policy_4610000.csv',
            table_type="accounting_policies",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='ACCOUNTING POLICIES')
    else:
        processed_wb.remove(processed_wb['ACCOUNTING POLICIES'])

    if id_availability_dict['4611100a']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_credit_currency_4611100a.csv',
            table_type="credit_by_currency",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='CREDIT BY CURRENCY')
    else:
        processed_wb.remove(processed_wb['CREDIT BY CURRENCY'])

    if id_availability_dict['4612100a']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_credit_type_4612100a.csv',
            table_type="credit_by_type",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=ap_df,
                                                workbook=processed_wb,
                                                sheet_name='CREDIT BY TYPE')
    else:
        processed_wb.remove(processed_wb['CREDIT BY TYPE'])

    if id_availability_dict['4613100a']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_credit_sector_4613100a.csv',
            table_type="credit_by_sector",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=ap_df,
                                                workbook=processed_wb,
                                                sheet_name='CREDIT BY SECTOR')
    else:
        processed_wb.remove(processed_wb['CREDIT BY SECTOR'])

    if id_availability_dict['4614100']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_credit_other_4614100.csv',
            table_type="credit_other",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='CREDIT OTHER INFORMATION')
    else:
        processed_wb.remove(processed_wb['CREDIT OTHER INFORMATION'])

    if id_availability_dict['4622100']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_savings_4622100.csv',
            table_type="savings_breakdown",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=ap_df,
                                                workbook=processed_wb,
                                                sheet_name='SAVINGS BREAKDOWN')
    else:
        processed_wb.remove(processed_wb['SAVINGS BREAKDOWN'])

    if id_availability_dict['4621100']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_giro_4621100.csv',
            table_type="giro_breakdown",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(df=ap_df,
                                                workbook=processed_wb,
                                                sheet_name='GIRO BREAKDOWN')
    else:
        processed_wb.remove(processed_wb['GIRO BREAKDOWN'])

    if id_availability_dict['4623100']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_deposito_4623100.csv',
            table_type="deposito_breakdown",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='TIME DEPOSITS BREAKDOWN')
    else:
        processed_wb.remove(processed_wb['TIME DEPOSITS BREAKDOWN'])

    if id_availability_dict['4624100']:
        ap_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_deposit_interest_4624100.csv',
            table_type="deposit_interest",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=ap_df, workbook=processed_wb, sheet_name='DEPOSIT INTEREST RATE')
    else:
        processed_wb.remove(processed_wb['DEPOSIT INTEREST RATE'])

    if id_availability_dict['4631100']:
        interest_rev_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_interest_revenue_4631100.csv',
            table_type="interest_revenue",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=interest_rev_df,
            workbook=processed_wb,
            sheet_name='INTEREST REVENUE BREAKDOWN')

        if quarterly:
            interest_rev_quarterly_df = convert_cumulative_to_quarterly(
                interest_rev_df)

            processed_wb = export_to_excel_template(
                df=interest_rev_quarterly_df,
                workbook=processed_wb,
                sheet_name='INTEREST REVENUE BREAKDOWN QoQ')
        else:
            processed_wb.remove(processed_wb['INTEREST REVENUE BREAKDOWN QoQ'])
    else:
        processed_wb.remove(processed_wb['INTEREST REVENUE BREAKDOWN'])
        processed_wb.remove(processed_wb['INTEREST REVENUE BREAKDOWN QoQ'])

    if id_availability_dict['4632100']:
        interest_expense_df = get_finance_financial_table(
            combined_df=df,
            name_mapping_path='./schema/finance_interest_expense_4632100.csv',
            table_type="interest_expense",
            convert_billion_idr=True,
            convert_million_usd=True)

        processed_wb = export_to_excel_template(
            df=interest_expense_df,
            workbook=processed_wb,
            sheet_name='INTEREST EXPENSE BREAKDOWN')

        if quarterly:
            interest_expense_quarterly_df = convert_cumulative_to_quarterly(
                interest_expense_df)

            processed_wb = export_to_excel_template(
                df=interest_expense_quarterly_df,
                workbook=processed_wb,
                sheet_name='INTEREST EXPENSE BREAKDOWN QoQ')
        else:
            processed_wb.remove(processed_wb['INTEREST EXPENSE BREAKDOWN QoQ'])
            
    else:
        processed_wb.remove(processed_wb['INTEREST EXPENSE BREAKDOWN'])
        processed_wb.remove(processed_wb['INTEREST EXPENSE BREAKDOWN QoQ'])

    if quarterly:
        processed_wb.save(
            f"./result/{stock_code}_Financial_Statement_Quarterly.xlsx")
    else:
        processed_wb.save(f"./result/{stock_code}_Financial_Statement.xlsx")


def get_and_export_financial_report(stock_code: str,
                                    years: list,
                                    quarterly: bool,
                                    convert_billion_idr: bool,
                                    convert_million_usd: bool):
    if quarterly:
        print(
            f"Getting {stock_code} quarterly financial report for period: {years}")
    else:
        print(
            f"Getting {stock_code} annual financial report for period: {years}")

    combined_df = download_and_combine_xbrl(stock_code=stock_code,
                                            years=years,
                                            quarterly=quarterly)

    main_industry = combined_df[combined_df.name ==
                                'EntityMainIndustry'].value.unique()[0]
    print(f"Industry is {main_industry}")

    id_availability_dict = get_id_availability(combined_df,
                                               main_industry=main_industry)

    if id_availability_dict is None:
        print("Industry is currently not supported")
        return None

    if "General" in main_industry:
        export_general_report_to_excel(
            stock_code=stock_code,
            df=combined_df,
            workbook_path='./template/General_Company_Template.xlsx',
            id_availability_dict=id_availability_dict,
            quarterly=quarterly,
            convert_billion_idr=convert_billion_idr,
            convert_million_usd=convert_million_usd)
    elif "Financial" in main_industry:
        export_finance_report_to_excel(
            stock_code=stock_code,
            df=combined_df,
            workbook_path='./template/Financial_Company_Template.xlsx',
            id_availability_dict=id_availability_dict,
            quarterly=quarterly,
            convert_billion_idr=True,
            convert_million_usd=True)
